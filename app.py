
import streamlit as st
import pandas as pd
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
from copy import copy
from pathlib import Path
import os

st.set_page_config(page_title="엑셀 행 재정렬 안전 비교 (전체열 + 색상)", layout="wide")
st.title("📘 엑셀 행 재정렬 안전 비교 (전체열 + 색상)")
st.caption("기준 파일과 비교 파일을 선택하면, 행 순서가 달라도 전체 열에서 **값 변경**과 **배경색(채우기) 변경**을 잡아냅니다.")

# 대용량 파일 안내
with st.expander("ℹ️ 사용 안내", expanded=False):
    st.info("""
    **권장 사항:**
    - 행 개수: 10,000개 이하 (초과 시 자동 제한)
    - 열 개수: 100개 이하 (초과 시 자동 제한)
    - 파일 크기: 50MB 이하
    
    **대용량 파일 처리:**
    - 10,000행 초과 시 처음 10,000행만 처리됩니다.
    - 100열 초과 시 처음 100열만 처리됩니다.
    - 메모리 부족 시 파일을 분할하여 처리하세요.
    """)

# ----------------------- 셀 스타일 복사 -----------------------
def copy_cell_style(source_cell, target_cell):
    """
    원본 셀의 스타일을 대상 셀로 복사합니다.
    """
    try:
        if source_cell.has_style:
            # 폰트 복사
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            
            # 채우기(배경색) 복사
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            
            # 테두리 복사
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            
            # 정렬 복사
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            
            # 숫자 형식 복사
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            
            # 보호 복사
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
    except Exception as e:
        pass  # 스타일 복사 실패는 무시

def copy_row_with_style(source_ws, target_ws, source_row_idx, target_row_idx, max_col):
    """
    원본 워크시트의 특정 행을 대상 워크시트로 스타일 포함하여 복사합니다.
    """
    try:
        for col in range(1, max_col + 1):
            source_cell = source_ws.cell(row=source_row_idx, column=col)
            target_cell = target_ws.cell(row=target_row_idx, column=col)
            
            # 값 복사
            target_cell.value = source_cell.value
            
            # 스타일 복사
            copy_cell_style(source_cell, target_cell)
        
        # 행 높이 복사
        if source_ws.row_dimensions[source_row_idx].height:
            target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[source_row_idx].height
    except Exception as e:
        pass  # 행 복사 실패는 무시

def copy_column_widths(source_ws, target_ws):
    """
    열 너비를 복사합니다.
    """
    try:
        for col_letter in source_ws.column_dimensions:
            if source_ws.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
    except Exception as e:
        pass

def copy_entire_sheet(source_ws, target_ws):
    """
    시트 전체를 스타일 포함하여 복사합니다.
    """
    try:
        max_row = source_ws.max_row
        max_col = source_ws.max_column
        
        # 모든 셀 복사
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=row, column=col)
                
                # 값 복사
                target_cell.value = source_cell.value
                
                # 스타일 복사
                copy_cell_style(source_cell, target_cell)
        
        # 열 너비 복사
        copy_column_widths(source_ws, target_ws)
        
        # 행 높이 복사
        for row_idx in source_ws.row_dimensions:
            if source_ws.row_dimensions[row_idx].height:
                target_ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height
    except Exception as e:
        st.warning(f"시트 복사 중 일부 오류 발생: {e}")

# ----------------------- 색상/채우기 라벨링 -----------------------
def _fill_is_nonempty(fill) -> bool:
    if fill is None:
        return False
    pt = getattr(fill, "patternType", None)
    if not pt or str(pt).lower() == "none":
        return False
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return True
    if getattr(fg, "rgb", None) or getattr(fg, "indexed", None) is not None or getattr(fg, "theme", None) is not None:
        return True
    return True

def _color_hex_from_fg(fg) -> str | None:
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    if isinstance(rgb, str):
        s = rgb.replace("#", "").upper()
        if len(s) == 8:
            s = s[2:]
        if len(s) == 6:
            return "#" + s
    idx = getattr(fg, "indexed", None)
    if idx is not None:
        mapping = {1:"#000000", 2:"#FFFFFF", 6:"#FFFF00"}
        return mapping.get(idx, f"indexed-{idx}")
    return None

def fill_to_label(fill) -> str:
    if fill is None:
        return "No Fill"
    pt = getattr(fill, "patternType", None)
    if not pt or str(pt).lower() == "none":
        return "No Fill"
    fg = getattr(fill, "fgColor", None)
    hx = _color_hex_from_fg(fg)
    if hx is None:
        return "Fill"
    friendly = {
        "#FFFFFF":"White",
        "#000000":"Black",
        # Yellow shades
        "#FFFF00":"Yellow",
        "#FFF2CC":"Light Yellow",
        "#FFD966":"Gold",
        "#FFEB9C":"Light Yellow 2",
        "#FFFF99":"Light Yellow (Alt)",
        "#FFFFCC":"Pale Yellow",
        # Red shades
        "#FF0000":"Red",
        "#FFC7CE":"Light Red",
        "#FFCCCC":"Pale Red",
        "#FF6666":"Light Red 2",
        # Green shades
        "#00FF00":"Green",
        "#00B050":"Dark Green",
        "#92D050":"Light Green",
        "#C6E0B4":"Pale Green",
        "#E2EFDA":"Very Light Green",
        # Blue shades
        "#0000FF":"Blue",
        "#00B0F0":"Light Blue",
        "#BDD7EE":"Pale Blue",
        "#DDEBF7":"Very Light Blue",
        # Orange shades
        "#FFA500":"Orange",
        "#F8CBAD":"Light Orange",
        "#FFC000":"Dark Orange",
        # Purple shades
        "#7030A0":"Purple",
        "#B4A7D6":"Light Purple",
        # Gray shades
        "#D9D9D9":"Light Gray",
        "#BFBFBF":"Gray",
        "#808080":"Dark Gray",
    }.get(hx)
    return friendly or hx

# ----------------------- 범위(행/열) 계산 -----------------------
def compute_used_bounds(ws, max_rows_limit=10000, max_cols_limit=100):
    """
    실제 사용된 행/열 범위를 계산 (대용량 파일 대응)
    """
    try:
        # 제한 적용
        max_possible_r = min(ws.max_row, max_rows_limit)
        max_possible_c = min(ws.max_column, max_cols_limit)
        
        max_r, max_c = 0, 0
        
        # 역순으로 검색하여 최적화
        for r in range(max_possible_r, 0, -1):
            row_has_any = False
            for c in range(1, max_possible_c + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    if (cell.value not in (None, "")) or _fill_is_nonempty(cell.fill):
                        row_has_any = True
                        if c > max_c:
                            max_c = c
                except Exception:
                    continue
            if row_has_any:
                max_r = r
                break
        
        # 최대 열 확인
        if max_r > 0 and max_c == 0:
            for r in range(1, min(max_r + 1, 100)):  # 샘플링
                for c in range(1, max_possible_c + 1):
                    try:
                        cell = ws.cell(row=r, column=c)
                        if (cell.value not in (None, "")) or _fill_is_nonempty(cell.fill):
                            if c > max_c:
                                max_c = c
                    except Exception:
                        continue
        
        if max_r == 0:
            max_r = min(ws.max_row, max_rows_limit)
        if max_c == 0:
            max_c = min(ws.max_column, max_cols_limit)
        
        return max_r, max_c
    except Exception as e:
        st.warning(f"범위 계산 중 오류 발생, 기본값 사용: {e}")
        return min(ws.max_row, max_rows_limit), min(ws.max_column, max_cols_limit)

# ----------------------- 정규화 -----------------------
def normalize_value(v, trim_spaces=True, case_sensitive=True):
    if isinstance(v, str):
        s = v.strip() if trim_spaces else v
        return s if case_sensitive else s.lower()
    return v

# ----------------------- 시트 읽기 -----------------------
def read_sheet_values_and_fills(file, sheet_name=None, trim_spaces=True, case_sensitive=True):
    """
    엑셀 시트의 값과 채우기 정보를 읽어옵니다.
    """
    wb = None
    try:
        # read_only=False로 열어야 스타일 정보를 읽을 수 있음
        wb = load_workbook(file, data_only=True, read_only=False)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if ws is None:
            raise ValueError("시트를 찾을 수 없습니다.")
        
        # 대용량 파일 경고
        if ws.max_row > 10000:
            st.warning(f"⚠️ 파일에 {ws.max_row}개의 행이 있습니다. 처음 10,000개 행만 처리합니다.")
        if ws.max_column > 100:
            st.warning(f"⚠️ 파일에 {ws.max_column}개의 열이 있습니다. 처음 100개 열만 처리합니다.")
        
        max_r, max_c = compute_used_bounds(ws)
        
        if max_r == 0 or max_c == 0:
            return [], {}, []
        
        cols = [get_column_letter(c) for c in range(1, max_c + 1)]

        rows = []
        fills = {}
        
        for r in range(1, max_r + 1):
            try:
                orig = {}
                norm = {}
                empty_all = True
                
                for c in range(1, max_c + 1):
                    try:
                        cell = ws.cell(row=r, column=c)
                        v = cell.value
                        col = get_column_letter(c)
                        orig[col] = v
                        norm[col] = normalize_value(v, trim_spaces, case_sensitive)
                        
                        # 채우기 정보
                        try:
                            fills[(r, c)] = fill_to_label(cell.fill)
                        except Exception:
                            fills[(r, c)] = "No Fill"
                        
                        if (v not in (None, "")) or _fill_is_nonempty(cell.fill):
                            empty_all = False
                    except Exception as e:
                        # 개별 셀 오류는 무시
                        col = get_column_letter(c)
                        orig[col] = None
                        norm[col] = None
                        fills[(r, c)] = "No Fill"
                
                if not empty_all:
                    rows.append({"_row": r, "orig": orig, "norm": norm})
            except Exception as e:
                st.warning(f"행 {r} 처리 중 오류 발생, 건너뜀: {e}")
                continue
        
        return rows, fills, cols
    
    except Exception as e:
        st.error(f"파일 읽기 실패: {e}")
        raise
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

# ----------------------- 페어링 -----------------------
def row_tuple(norm_row, columns):
    return tuple(norm_row.get(col) for col in columns)

def best_pairing(new_rows, old_rows, columns):
    """
    최적 페어링 알고리즘 (대용량 데이터 대응)
    """
    candidates = []
    
    # 대용량 데이터 처리
    max_pairs_to_check = 50000  # 최대 확인할 페어 수
    
    try:
        for i, o in enumerate(old_rows):
            for j, n in enumerate(new_rows):
                # 너무 많은 페어는 건너뜀
                if len(candidates) > max_pairs_to_check:
                    break
                
                try:
                    eq = sum(1 for col in columns if o["norm"].get(col) == n["norm"].get(col))
                    if eq > 0:
                        candidates.append((eq, i, j))
                except Exception:
                    continue
            
            if len(candidates) > max_pairs_to_check:
                break
        
        if len(candidates) > max_pairs_to_check:
            st.warning(f"⚠️ 페어링 후보가 너무 많습니다 ({len(candidates)}개). 상위 {max_pairs_to_check}개만 처리합니다.")
            candidates = candidates[:max_pairs_to_check]
        
        candidates.sort(reverse=True)
        used_old, used_new = set(), set()
        pairs = []
        
        for eq, i, j in candidates:
            if i in used_old or j in used_new:
                continue
            pairs.append((i, j, eq))
            used_old.add(i)
            used_new.add(j)
        
        leftover_old = [i for i in range(len(old_rows)) if i not in used_old]
        leftover_new = [j for j in range(len(new_rows)) if j not in used_new]
        
        return pairs, leftover_old, leftover_new
    
    except Exception as e:
        st.error(f"페어링 중 오류 발생: {e}")
        return [], list(range(len(old_rows))), list(range(len(new_rows)))

# ----------------------- 변경 레코드 -----------------------
def truncate_value(val, max_len=50):
    """값이 너무 길면 잘라냅니다."""
    if val is None:
        return ""
    s = str(val)
    if len(s) > max_len:
        return s[:max_len] + "..."
    return s

def build_diff_record(old_row, new_row, old_fills, new_fills, columns):
    """변경 사항을 기록합니다."""
    changes = []
    try:
        for idx, col in enumerate(columns, start=1):
            try:
                r_old = old_row["_row"]
                r_new = new_row["_row"]
                ov = old_row["orig"].get(col)
                nv = new_row["orig"].get(col)
                value_changed = old_row["norm"].get(col) != new_row["norm"].get(col)

                ofill = old_fills.get((r_old, idx), "No Fill")
                nfill = new_fills.get((r_new, idx), "No Fill")
                fill_changed = ofill != nfill

                if value_changed or fill_changed:
                    # 값을 잘라서 표시
                    ov_str = truncate_value(ov, 30)
                    nv_str = truncate_value(nv, 30)
                    
                    if value_changed and fill_changed:
                        changes.append(f"{col}열 값 '{ov_str}'→'{nv_str}', 색 '{ofill}'→'{nfill}'")
                    elif value_changed:
                        changes.append(f"{col}열 값 '{ov_str}'→'{nv_str}'")
                    elif fill_changed:
                        changes.append(f"{col}열 색 '{ofill}'→'{nfill}'")
            except Exception as e:
                changes.append(f"{col}열 처리 오류")
                continue
        
        # 변경 사항이 너무 많으면 요약
        if len(changes) > 10:
            msg = f"{len(changes)}개 열 변경됨 (처음 10개: " + "; ".join(changes[:10]) + "...)"
        else:
            msg = "; ".join(changes) if changes else "변경 없음"
        
        return {
            "기준행": old_row["_row"],
            "비교행": new_row["_row"],
            "변경요약": msg
        }
    except Exception as e:
        return {
            "기준행": old_row.get("_row", "?"),
            "비교행": new_row.get("_row", "?"),
            "변경요약": f"처리 오류: {str(e)[:50]}"
        }

# ----------------------- 로컬 폴더에서 파일 가져오기 -----------------------
def get_excel_files_in_folder(folder_path):
    """폴더 내의 모든 엑셀 파일 목록 반환"""
    try:
        if not folder_path:
            return []
        
        # 경로 정규화
        folder_path = os.path.normpath(folder_path)
        
        if not os.path.exists(folder_path):
            return []
        
        if not os.path.isdir(folder_path):
            return []
        
        path = Path(folder_path)
        excel_files = []
        
        try:
            excel_files = list(path.glob("*.xlsx")) + list(path.glob("*.xls"))
        except Exception as e:
            st.warning(f"파일 검색 중 오류: {e}")
            return []
        
        # 임시 파일 및 숨김 파일 제외
        excel_files = [f for f in excel_files if not f.name.startswith("~$") and not f.name.startswith(".")]
        
        return sorted([f.name for f in excel_files])
    except Exception as e:
        st.error(f"폴더 읽기 오류: {e}")
        return []

# ----------------------- UI -----------------------
with st.expander("⚙️ 설정", expanded=True):
    col_opt1, col_opt2 = st.columns(2)
    with col_opt1:
        trim_spaces = st.checkbox("앞뒤 공백 무시", value=True)
        case_sensitive = st.checkbox("대소문자 구분", value=True)
    with col_opt2:
        # 파일 입력 방식 선택
        input_mode = st.radio("파일 입력 방식", ["로컬 폴더", "파일 업로드"], horizontal=True)

st.subheader("1️⃣ 기준(이전) 파일 선택")

if input_mode == "로컬 폴더":
    # 현재 작업 디렉토리를 기본값으로 사용
    default_folder = os.getcwd()
    folder_path = st.text_input("📁 폴더 경로", value=default_folder, help="엑셀 파일이 있는 폴더 경로를 입력하세요")
    
    if folder_path and os.path.exists(folder_path):
        excel_files = get_excel_files_in_folder(folder_path)
        
        if excel_files:
            c1, c2 = st.columns(2)
            with c1:
                selected_old_file = st.selectbox("기준 파일 선택", options=excel_files, key="old_file_select")
                file_old = os.path.join(folder_path, selected_old_file) if selected_old_file else None
            with c2:
                sheet_old = None
                if file_old:
                    wb = None
                    try:
                        wb = load_workbook(file_old, read_only=True, data_only=True)
                        if wb and wb.sheetnames:
                            sheet_old = st.selectbox("시트 선택(기준)", options=wb.sheetnames, index=0, key="old_sheet")
                        else:
                            st.error("시트를 찾을 수 없습니다.")
                    except Exception as e:
                        st.error(f"기준 파일 시트 읽기 실패: {e}")
                    finally:
                        if wb:
                            try:
                                wb.close()
                            except Exception:
                                pass
        else:
            st.warning("⚠️ 선택한 폴더에 엑셀 파일이 없습니다.")
            file_old = None
            sheet_old = None
    else:
        st.warning("⚠️ 유효한 폴더 경로를 입력하세요.")
        file_old = None
        sheet_old = None
else:
    # 파일 업로드 방식
    c1, c2 = st.columns(2)
    with c1:
        file_old = st.file_uploader("기준 엑셀 파일", type=["xlsx"], key="old_allcols")
    with c2:
        sheet_old = None
        if file_old:
            wb = None
            try:
                wb = load_workbook(file_old, read_only=True, data_only=True)
                if wb and wb.sheetnames:
                    sheet_old = st.selectbox("시트 선택(기준)", options=wb.sheetnames, index=0)
                else:
                    st.error("시트를 찾을 수 없습니다.")
            except Exception as e:
                st.error(f"기준 파일 시트 읽기 실패: {e}")
            finally:
                if wb:
                    try:
                        wb.close()
                    except Exception:
                        pass

if st.button("✅ 기준 데이터 저장", type="primary", disabled=not (file_old and sheet_old)):
    try:
        with st.spinner("기준 파일을 읽는 중..."):
            old_rows, old_fills, cols = read_sheet_values_and_fills(file_old, sheet_old, trim_spaces, case_sensitive)
            
            if not old_rows:
                st.error("❌ 기준 파일에 데이터가 없습니다.")
            else:
                st.session_state["old_rows"] = old_rows
                st.session_state["old_fills"] = old_fills
                st.session_state["columns"] = cols
                st.session_state["trim_spaces"] = trim_spaces
                st.session_state["case_sensitive"] = case_sensitive
                
                # 원본 파일 정보 저장 (스타일 복사용)
                st.session_state["old_file_path"] = file_old
                st.session_state["old_sheet_name"] = sheet_old

                multiset = Counter([row_tuple(r["norm"], cols) for r in old_rows])
                mapping = defaultdict(list)
                for idx, r in enumerate(old_rows):
                    mapping[row_tuple(r["norm"], cols)].append(idx)

                st.session_state["old_rows_norm_multiset"] = multiset
                st.session_state["old_rows_by_tuple_indices"] = mapping
                st.success(f"✅ 기준 데이터 저장 완료: {len(old_rows)} 행, 사용 열: {len(cols)}개 ({cols[0]}~{cols[-1]})")
    except Exception as e:
        st.error(f"❌ 기준 파일 처리 중 오류 발생")
        st.exception(e)

st.subheader("2️⃣ 비교(이후) 파일 선택")

if input_mode == "로컬 폴더":
    # 같은 폴더에서 비교 파일 선택
    if folder_path and os.path.exists(folder_path):
        excel_files = get_excel_files_in_folder(folder_path)
        
        if excel_files:
            c3, c4 = st.columns(2)
            with c3:
                selected_new_file = st.selectbox("비교 파일 선택", options=excel_files, key="new_file_select")
                file_new = os.path.join(folder_path, selected_new_file) if selected_new_file else None
            with c4:
                sheet_new = None
                if file_new:
                    wb2 = None
                    try:
                        wb2 = load_workbook(file_new, read_only=True, data_only=True)
                        if wb2 and wb2.sheetnames:
                            sheet_new = st.selectbox("시트 선택(비교)", options=wb2.sheetnames, index=0, key="new_sheet")
                        else:
                            st.error("시트를 찾을 수 없습니다.")
                    except Exception as e:
                        st.error(f"비교 파일 시트 읽기 실패: {e}")
                    finally:
                        if wb2:
                            try:
                                wb2.close()
                            except Exception:
                                pass
        else:
            file_new = None
            sheet_new = None
    else:
        file_new = None
        sheet_new = None
else:
    # 파일 업로드 방식
    c3, c4 = st.columns(2)
    with c3:
        file_new = st.file_uploader("비교 엑셀 파일", type=["xlsx"], key="new_allcols")
    with c4:
        sheet_new = None
        if file_new:
            wb2 = None
            try:
                wb2 = load_workbook(file_new, read_only=True, data_only=True)
                if wb2 and wb2.sheetnames:
                    sheet_new = st.selectbox("시트 선택(비교)", options=wb2.sheetnames, index=0)
                else:
                    st.error("시트를 찾을 수 없습니다.")
            except Exception as e:
                st.error(f"비교 파일 시트 읽기 실패: {e}")
            finally:
                if wb2:
                    try:
                        wb2.close()
                    except Exception:
                        pass

if st.button("🔍 변경 사항 분석 실행", type="primary",
             disabled=not (file_new and sheet_new and ("old_rows" in st.session_state))):
    try:
        # 저장된 설정값 사용
        old_rows = st.session_state["old_rows"]
        old_fills = st.session_state["old_fills"]
        columns_old = st.session_state["columns"]
        old_multiset = st.session_state["old_rows_norm_multiset"]
        old_tuple_to_indices = st.session_state["old_rows_by_tuple_indices"]
        saved_trim_spaces = st.session_state.get("trim_spaces", trim_spaces)
        saved_case_sensitive = st.session_state.get("case_sensitive", case_sensitive)
        
        # 비교 파일 정보 저장 (스타일 복사용)
        st.session_state["new_file_path"] = file_new
        st.session_state["new_sheet_name"] = sheet_new

        # 진행 상황 표시
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("📖 비교 파일을 읽는 중...")
        progress_bar.progress(10)
        
        new_rows, new_fills, cols_new = read_sheet_values_and_fills(
            file_new, sheet_new, saved_trim_spaces, saved_case_sensitive
        )
        
        if not new_rows:
            st.error("❌ 비교 파일에 데이터가 없습니다.")
            progress_bar.empty()
            status_text.empty()
        else:
            progress_bar.progress(20)
            
            # 열 범위: 기준/비교 중 더 넓은 범위를 사용 (기존 columns_old는 유지)
            all_columns = list(set(columns_old + cols_new))
            all_columns.sort(key=lambda x: (len(x), x))  # A, B, ... Z, AA, AB ...
            columns = all_columns

            status_text.text("🔄 동일한 행 매칭 중...")
            progress_bar.progress(30)
            
            remaining_old_indices = set(range(len(old_rows)))
            remaining_new_indices = set(range(len(new_rows)))

            exact_pairs = []
            temp_multiset = old_multiset.copy()
            temp_tuple_to_indices = {k: v.copy() for k, v in old_tuple_to_indices.items()}

            for j, nr in enumerate(new_rows):
                t = row_tuple(nr["norm"], columns)
                if temp_multiset.get(t, 0) > 0:
                    i = temp_tuple_to_indices[t].pop(0)
                    temp_multiset[t] -= 1
                    exact_pairs.append((i, j))
                    remaining_old_indices.discard(i)
                    remaining_new_indices.discard(j)

            progress_bar.progress(50)
            status_text.text("🔍 변경된 행 매칭 중...")
            
            old_left = [old_rows[i] for i in sorted(remaining_old_indices)]
            new_left = [new_rows[j] for j in sorted(remaining_new_indices)]
            pairs, leftover_old_idx, leftover_new_idx = best_pairing(new_left, old_left, columns)

            progress_bar.progress(60)
            status_text.text("📊 변경 내역 생성 중...")
            
            best_pairs = []
            sorted_old_left = sorted(remaining_old_indices)
            sorted_new_left = sorted(remaining_new_indices)
            for eq, i, j in sorted([(p[2], p[0], p[1]) for p in pairs], reverse=True):
                old_idx_global = sorted_old_left[i]
                new_idx_global = sorted_new_left[j]
                best_pairs.append((old_idx_global, new_idx_global, eq))

            unchanged_records = [{
                "기준행": old_rows[i]["_row"],
                "비교행": new_rows[j]["_row"],
                "상태": "동일(재정렬만)"
            } for i, j in exact_pairs]

            progress_bar.progress(70)
            
            changes_records = []
            for i, j, eq in best_pairs:
                rec = build_diff_record(old_rows[i], new_rows[j], old_fills, new_fills, columns)
                rec["일치열수"] = eq
                rec["상태"] = "변경"
                changes_records.append(rec)

            progress_bar.progress(80)
            
            used_old = set([i for i, _, _ in best_pairs] + [i for i, _ in exact_pairs])
            used_new = set([j for _, j, _ in best_pairs] + [j for _, j in exact_pairs])

            removed_records = [{"기준행": old_rows[i]["_row"], "상태": "제거됨"} for i in range(len(old_rows)) if i not in used_old]
            added_records = [{"비교행": new_rows[j]["_row"], "상태": "추가됨"} for j in range(len(new_rows)) if j not in used_new]

            progress_bar.progress(90)
            status_text.text("✨ 결과 정리 중...")
            
            df_unchanged = pd.DataFrame(unchanged_records)
            df_changes = pd.DataFrame(changes_records, columns=["기준행","비교행","일치열수","변경요약","상태"])
            df_removed = pd.DataFrame(removed_records)
            df_added = pd.DataFrame(added_records)
            
            # 세션에 저장
            st.session_state["df_unchanged"] = df_unchanged
            st.session_state["df_changes"] = df_changes
            st.session_state["df_removed"] = df_removed
            st.session_state["df_added"] = df_added
            
            progress_bar.progress(100)
            status_text.text("✅ 분석 완료!")
            
            st.success(f"✅ 분석 완료: 동일(재정렬만) {len(df_unchanged)}건, 변경 {len(df_changes)}건, 제거 {len(df_removed)}건, 추가 {len(df_added)}건")
            
            progress_bar.empty()
            status_text.empty()
    
    except Exception as e:
        if 'progress_bar' in locals():
            progress_bar.empty()
        if 'status_text' in locals():
            status_text.empty()
        st.error("❌ 분석 중 오류가 발생했습니다.")
        st.exception(e)

# ----------------------- 결과 표시 -----------------------
if "df_unchanged" in st.session_state:
    st.divider()
    st.subheader("📊 분석 결과")
    
    df_unchanged = st.session_state["df_unchanged"]
    df_changes = st.session_state["df_changes"]
    df_removed = st.session_state["df_removed"]
    df_added = st.session_state["df_added"]
    
    # 필터링 옵션
    with st.expander("🔍 결과 필터링", expanded=False):
        show_unchanged = st.checkbox("동일(재정렬만) 표시", value=True)
        show_changes = st.checkbox("변경 사항 표시", value=True)
        show_removed = st.checkbox("제거된 행 표시", value=True)
        show_added = st.checkbox("추가된 행 표시", value=True)
        
        if show_changes and not df_changes.empty:
            search_text = st.text_input("🔎 변경 내용 검색", placeholder="검색어를 입력하세요 (변경요약에서 검색)")
    
    # 동일(재정렬만)
    if show_unchanged:
        st.write("### ✅ 동일(재정렬만)")
        if not df_unchanged.empty:
            st.dataframe(df_unchanged, use_container_width=True, hide_index=True)
        else:
            st.info("동일한 행이 없습니다.")
    
    # 변경
    if show_changes:
        st.write("### 🔄 변경 (값/색상)")
        if not df_changes.empty:
            df_to_show = df_changes.copy()
            if 'search_text' in locals() and search_text:
                df_to_show = df_to_show[df_to_show["변경요약"].str.contains(search_text, case=False, na=False)]
                st.caption(f"검색 결과: {len(df_to_show)}건")
            st.dataframe(df_to_show, use_container_width=True, hide_index=True)
        else:
            st.info("변경된 행이 없습니다.")
    
    # 제거됨
    if show_removed:
        st.write("### ❌ 제거됨 (기준에는 있었으나 비교에는 없음)")
        if not df_removed.empty:
            st.dataframe(df_removed, use_container_width=True, hide_index=True)
        else:
            st.info("제거된 행이 없습니다.")
    
    # 추가됨
    if show_added:
        st.write("### ➕ 추가됨 (비교에는 있으나 기준에는 없음)")
        if not df_added.empty:
            st.dataframe(df_added, use_container_width=True, hide_index=True)
        else:
            st.info("추가된 행이 없습니다.")

    # 다운로드 버튼
    st.divider()
    st.subheader("💾 결과 다운로드")
    
    from io import BytesIO
    
    def create_result_excel_with_styles():
        """
        실제 엑셀 셀과 스타일을 복사하여 결과 파일 생성
        Sheet1: 변경된 내용 (기준 행 + 비교 행)
        Sheet2: 추가된 내용 (비교 파일에서 복사)
        Sheet3: 삭제된 내용 (기준 파일에서 복사)
        Sheet4: 원본 기준 엑셀 전체
        """
        try:
            # 원본 파일 정보 가져오기
            old_file_path = st.session_state.get("old_file_path")
            old_sheet_name = st.session_state.get("old_sheet_name")
            new_file_path = st.session_state.get("new_file_path")
            new_sheet_name = st.session_state.get("new_sheet_name")
            
            if not old_file_path or not old_sheet_name:
                st.error("원본 파일 정보가 없습니다. 기준 데이터를 먼저 저장해주세요.")
                return None
            
            # 원본 워크북 열기
            wb_old = load_workbook(old_file_path)
            ws_old = wb_old[old_sheet_name]
            
            wb_new = None
            ws_new = None
            if new_file_path and new_sheet_name:
                wb_new = load_workbook(new_file_path)
                ws_new = wb_new[new_sheet_name]
            
            # 결과 워크북 생성
            result_wb = Workbook()
            result_wb.remove(result_wb.active)  # 기본 시트 제거
            
            # 최대 열 수 계산
            max_col = ws_old.max_column
            if ws_new:
                max_col = max(max_col, ws_new.max_column)
            
            # Sheet1: 변경된 내용
            if not df_changes.empty:
                ws_changes = result_wb.create_sheet("변경된내용")
                current_row = 1
                
                # 헤더 추가
                ws_changes.cell(row=current_row, column=1, value="[기준 파일]")
                current_row += 1
                
                for idx, row in df_changes.iterrows():
                    old_row_num = row["기준행"]
                    new_row_num = row["비교행"]
                    
                    # 구분선
                    ws_changes.cell(row=current_row, column=1, value=f"--- 행 {old_row_num} → {new_row_num} ---")
                    current_row += 1
                    
                    # 기준 파일의 행 복사
                    ws_changes.cell(row=current_row, column=1, value="[변경 전]")
                    current_row += 1
                    copy_row_with_style(ws_old, ws_changes, old_row_num, current_row, max_col)
                    current_row += 1
                    
                    # 비교 파일의 행 복사
                    if ws_new:
                        ws_changes.cell(row=current_row, column=1, value="[변경 후]")
                        current_row += 1
                        copy_row_with_style(ws_new, ws_changes, new_row_num, current_row, max_col)
                        current_row += 1
                    
                    current_row += 1  # 빈 행 추가
                
                copy_column_widths(ws_old, ws_changes)
            
            # Sheet2: 추가된 내용
            if not df_added.empty and ws_new:
                ws_added = result_wb.create_sheet("추가된내용")
                current_row = 1
                
                for idx, row in df_added.iterrows():
                    new_row_num = row["비교행"]
                    copy_row_with_style(ws_new, ws_added, new_row_num, current_row, max_col)
                    current_row += 1
                
                copy_column_widths(ws_new, ws_added)
            
            # Sheet3: 삭제된 내용
            if not df_removed.empty:
                ws_removed = result_wb.create_sheet("삭제된내용")
                current_row = 1
                
                for idx, row in df_removed.iterrows():
                    old_row_num = row["기준행"]
                    copy_row_with_style(ws_old, ws_removed, old_row_num, current_row, max_col)
                    current_row += 1
                
                copy_column_widths(ws_old, ws_removed)
            
            # Sheet4: 원본 기준 엑셀 전체
            ws_original = result_wb.create_sheet("원본기준엑셀")
            copy_entire_sheet(ws_old, ws_original)
            
            # 워크북 저장
            bio = BytesIO()
            result_wb.save(bio)
            bio.seek(0)
            
            # 워크북 닫기
            wb_old.close()
            if wb_new:
                wb_new.close()
            result_wb.close()
            
            return bio.getvalue()
            
        except Exception as e:
            st.error(f"결과 파일 생성 중 오류: {e}")
            st.exception(e)
            return None
    
    # 스타일 포함 엑셀 다운로드
    st.info("💡 다운로드 파일에는 원본 엑셀의 **모든 색상과 스타일**이 포함됩니다.")
    
    try:
        with st.spinner("엑셀 파일 생성 중... (스타일 복사 중)"):
            result_data = create_result_excel_with_styles()
        
        if result_data:
            col_dl1, col_dl2 = st.columns(2)
            
            with col_dl1:
                st.download_button(
                    "📥 결과 다운로드 (원본 색상 포함)",
                    data=result_data,
                    file_name="excel_compare_with_styles.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            
            with col_dl2:
                st.success(f"""
                ✅ 다운로드 파일 구성:
                - Sheet1: 변경된 내용 ({len(df_changes)}건)
                - Sheet2: 추가된 내용 ({len(df_added)}건)
                - Sheet3: 삭제된 내용 ({len(df_removed)}건)
                - Sheet4: 원본 기준 엑셀 (전체)
                """)
        else:
            st.error("결과 파일 생성에 실패했습니다.")
    except Exception as e:
        st.error(f"결과 다운로드 준비 중 오류: {e}")
        st.exception(e)

st.divider()
st.info("💡 **사용 방법**: 기준 파일을 먼저 저장한 후, 비교 파일을 선택하여 분석을 실행하세요. 행 순서가 달라도 정확히 매칭하며, 모든 사용된 열(값/채우기 존재)을 자동 인식하여 비교합니다.")
